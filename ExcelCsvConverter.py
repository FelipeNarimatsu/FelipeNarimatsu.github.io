  
#! python3

# Chapter 14 Practice Excel to CSV Converter
# Converts all Excel file sheets in the current directory to CSV files

import os
import csv
import openpyxl

for excelFile in os.listdir('.'):
    if excelFile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excelFile)
        for sheetName in wb.sheetnames:
            sheet = wb.sheetnames
            csvFileName = open(excelFile + sheetName + '.csv', 'w', newline='')
            csvFile = csv.writer(csvFileName)
        
            for rowNum in range(1, 10):
                rowData = []
                for colNum in range(1, 10):
                    cellData = sheet.cell.internal_value(row=rowNum, column=colNum).value
                    rowData.append(cellData)

                csvFile.writerow(rowData)

            csvFileName.close()

       